import sqlite3, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
con = sqlite3.connect(r"C:\Users\ebmed\Documents\Car Sale Dataset\carsalesdatabase.db")
def typehelper(sheet, column_letter):
	for val in range(2, 500):
		if sheet[column_letter + str(val)].value is not None or sheet[column_letter + str(val)].value != 'NA':
			if type(sheet[column_letter + str(val)].value) == str:
				return 'text'
				break
			else:
				return 'real'
				break

			
def tablehelper(file_path):
	wb = openpyxl.load_workbook(file_path)
	localsheet = wb[wb.sheetnames[0]]
	x = 1
	tabledict = {}
	while localsheet[get_column_letter(x) + str(1)].value is not None:
		tabledict.update({localsheet[get_column_letter(x) + str(1)].value : typehelper(localsheet, get_column_letter(x))})
		x += 1
	return tabledict

def is_null_helper(sheet, number, num_columns):
        l_bool = True
        for val in range(1, num_columns + 1):
                if sheet[get_column_letter(val) + str(number)].value is not None:
                        pass
                else:
                        l_bool = False
        return l_bool

def insert_values(excel_file_path, table_name, num_columns):
	wb = openpyxl.load_workbook(excel_file_path)
	localsheet = wb[wb.sheetnames[0]]
	row_number = 2
	cur = con.cursor()
	insert_text = ' ('
	while is_null_helper(localsheet, row_number, num_columns):
		for column in range(1, num_columns + 1):
			if column == num_columns:
				if localsheet[get_column_letter(column) + str(row_number)].value == 'NA':
					insert_text += 'NULL )'
				elif type(localsheet[get_column_letter(column) + str(row_number)].value) == str:
					insert_text += "'" + localsheet[get_column_letter(column) + str(row_number)].value + "'" + " )"
				else:
					insert_text += str(localsheet[get_column_letter(column) + str(row_number)].value) + ' )'
			else:
				if localsheet[get_column_letter(column) + str(row_number)].value == 'NA':
					insert_text += 'NULL, '
				elif type(localsheet[get_column_letter(column) + str(row_number)].value) == str:
					insert_text += "'" + localsheet[get_column_letter(column) + str(row_number)].value + "'" + ", "
				else:
					insert_text += str(localsheet[get_column_letter(column) + str(row_number)].value) + ', '
		cur.execute("INSERT INTO " + table_name + " VALUES " + insert_text)
		insert_text = ' ('
		row_number += 1
	con.commit()


def tablecreator(excel_file_path):
	tabledict = tablehelper(excel_file_path)
	table_text = '('
	for key in tabledict.keys():
		if key == list(tabledict.keys())[-1]:
			table_text += key + ' ' + tabledict[key] + ')'
		else:
			table_text += key + ' ' + tabledict[key] + ', '
	cur = con.cursor()
	name_table = input('Please provide the table name here: ')
	cur.execute("CREATE TABLE " + name_table + table_text)
	con.commit()
	print("CREATE TABLE " + name_table + table_text)
	insert_values(excel_file_path, name_table, len(list(tabledict.keys())))

	
